# -*- coding: utf-8 -*-
"""Simplified case catalog: Type cau hoi | Cau hoi vi du cu the | Date check."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

# (type, example_question)
ROWS = [
("4-Key", "Vì sao DIG sai sóng sai ngành?"),
("4-Key", "SSI đang thuộc nhóm 4 Key nào?"),
("4-Key", "Lịch sử 4 Key của VCB từ 01/2026"),
("4-Key", "Cung cấp danh sách các mã đúng sóng đúng ngành ngày 24/08/2026"),
("4-Key", "Đánh giá 4 Key của SSI, VCB, HPG ngày 24/08/2026"),

("Mã", "Giá SSI hiện tại là bao nhiêu?"),
("Mã", "Giá SSI ngày 20/08/2026 là bao nhiêu?"),
("Mã", "Lịch sử mua bán SSI trong giai đoạn 2023-2025"),
("Mã", "Mã SSI bắt đầu mạnh từ khi nào?"),
("Mã", "Mã nào đạt chuẩn mã mạnh vào tháng 07/2026?"),
("Mã", "Giá SSI trong 20 phiên gần nhất là bao nhiêu?"),
("Mã", "Lập bảng thống kê SSI từ 2022 đến nay"),
("Mã", "Ngày 24/08/2026 cổ phiếu nào mạnh nhất dòng Ngân hàng?"),

("Sóng", "Ngành nào đang dẫn dắt hôm nay?"),
("Sóng", "Lộ trình các dòng dẫn sóng trong tháng 08/2026?"),
("Sóng", "Ngành chủ lực nào dẫn sóng vào 24/08/2026?"),
("Sóng", "Ngành Ngân hàng mất vai trò dẫn sóng khi nào?"),
("Định nghĩa", "(áp dụng ngầm khi trả lời các câu hỏi về ngành mạnh) — điều kiện SMDT 60-70% = có tiềm năng dẫn sóng"),
("Định nghĩa", "(áp dụng ngầm khi trả lời các câu hỏi về dẫn sóng) — ngưỡng SMDT=70% là \"đang dẫn sóng\""),
("Sóng", "Ngành Ngân hàng bắt đầu dẫn sóng từ khi nào?"),
("Sóng", "SSI đạt chuẩn ngành mạnh từ khi nào? (SSI thuộc ngành Chứng khoán)"),
("Sóng", "Ngành chủ lực nào dẫn sóng hôm nay?"),
("Sóng", "Sóng tháng 07/2026 ngành nào dẫn sóng?"),
("Ngành", "Ngành nào đạt chuẩn ngành mạnh vào năm 2026?"),
("Ngành", "Phân tích ngành Thép"),
("Ngành", "Sức mạnh dòng tiền ngành chủ lực vào 24/08/2026 thế nào?"),

("Chờ mua/bán", "Chờ mua ngày 24/08/2026 là bao nhiêu?"),

("Ngành", "Sức mạnh dòng tiền các ngành chủ lực ngày 24/08/2026"),
("Ngành", "SMDT ngành Ngân hàng từ tháng 01/2026 đến nay"),
("Ngành", "SMDT ngành Thép là bao nhiêu?"),
("Mã", "SMDT SSI là bao nhiêu?"),
("Mã", "Lập bảng thống kê mã giá giảm hôm nay mà SMDT tăng dần đều"),
("Mã", "SMDT SSI từ tháng 01/2026 đến nay"),
("Ngành", "SMDT các mã dòng Ngân hàng của ngày 24/08/2026"),
("Ngành", "SMDT các mã dòng Ngân hàng từ tháng 01/2026 đến nay"),
("Mã/Ngành", "SMDT SSI trong 10 phiên vừa qua"),
("Mã", "Các mã có SMDT tăng dần đều vào 24/08/2026"),
("Mã", "Lộ trình các mã suy yếu trong năm 2026"),
("Ngành", "Lộ trình các dòng chủ lực suy yếu trong tháng 07/2026"),
("Ngành", "Ngành nào có số mã có SMDT tăng dần đều nhiều nhất vào 24/08/2026?"),
("Ngành", "Danh sách các mã vượt 70% kèm giá của ngành Ngân hàng vào 24/08/2026"),

("Dòng tiền", "Dòng tiền ngành Ngân hàng bắt đầu đổ vào tháng 07/2026 khi nào?"),
("Dòng tiền", "Dòng tiền ngành Ngân hàng hiện nay thế nào?"),
("Dòng tiền", "Dòng tiền SSI hiện nay thế nào?"),
("Dòng tiền", "Dòng tiền SSI ngày 24/08/2026 là bao nhiêu?"),
("Dòng tiền", "Tín hiệu dòng tiền ACB từ 4/5 đến nay (ngày hiện tại giả định 2026-06-12)"),

("Mã", "Giá vốn trung bình và tỷ trọng nắm giữ của SSI hiện tại?"),
("Mã", "Tín hiệu mua bán gần nhất của SSI?"),
("Mã", "Chỉ số của SSI trong 15 phiên gần nhất?"),

("Sóng", "28/07 là sóng lớn hay sóng hồi?"),
("Sóng", "Chân sóng gần nhất là ngày nào?"),
("Sóng", "Phiên xác nhận tạo đáy gần nhất là ngày nào?"),
("Sóng", "Trong tháng 07/2026 có những phiên xác nhận tạo đáy nào?"),

("Sóng", "Hiệu suất cổ phiếu SSI khi ngành dẫn sóng?"),
("Sóng", "Hiệu suất các cổ phiếu ngành Ngân hàng vào tháng 07/2026 khi ngành dẫn sóng?"),
]

wb = Workbook()
ws = wb.active
ws.title = "Case Catalog"

headers = ["Type câu hỏi", "Câu hỏi ví dụ cụ thể", "Date check"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER
ws.freeze_panes = "A2"

for typ, question in ROWS:
    ws.append([typ, question, None])

for row_idx in range(2, len(ROWS) + 2):
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = CELL_FONT
        c.alignment = WRAP
        c.border = BORDER
    ws.cell(row=row_idx, column=3).number_format = "dd/mm/yyyy"

widths = [16, 70, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row_idx in range(2, len(ROWS) + 2):
    ws.row_dimensions[row_idx].height = 32

out_path = "case_catalog_simple.xlsx"
wb.save(out_path)
print("Saved", out_path, "rows:", len(ROWS))

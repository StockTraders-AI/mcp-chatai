# -*- coding: utf-8 -*-
"""Builds case_catalog.xlsx from the 11 rule files, one row per distinct
Guide block, with a concrete filled-in example question for any guide that
uses a placeholder like [ngành]/[date]/[ticker]."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
GROUP_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GROUP_FONT = Font(name=FONT_NAME, size=11, bold=True)
CELL_FONT = Font(name=FONT_NAME, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

# (source_file, guide_name, intent, api, args, example_question, notes)
ROWS = [
("Cau hoi ve danh gia 4 key co phieu.txt", "Phân tích cổ phiếu [ticker]",
 "Hỏi vì sao/tại sao/lý do/giải thích/chi tiết về 4-key/composite của 1 mã",
 "getStock4KeyEvaluation", "mode=single, ticker, date=YYYY-MM-DD, include_composite=true",
 "Vì sao DIG sai sóng sai ngành?",
 "Trả lời đủ 5 mục: Điểm Composite, Nhóm 4 Key + khuyến nghị, SMDT & Động lực, Phân kỳ, Bonus/Ghi chú. Không bịa số liệu ngoài tool."),

("Cau hoi ve danh gia 4 key co phieu.txt", "4 key đơn thuần [ticker]",
 "Hỏi trực tiếp: phân tích/đánh giá/trạng thái/key nào/thuộc nhóm nào/4 key",
 "getStock4KeyEvaluation", "mode=single, ticker, date=YYYY-MM-DD, include_composite=false",
 "SSI đang thuộc nhóm 4 Key nào?",
 "Chỉ trả lời ngắn gọn tên nhóm. Nếu hỏi \"có đúng sóng đúng ngành không\" → trả lời Có/Không + nhóm hiện tại."),

("Cau hoi ve danh gia 4 key co phieu.txt", "Lịch sử 4 key [ticker] từ [date]",
 "Hỏi lịch sử 4-key của 1 mã từ 1 ngày trở về sau",
 "getStock4KeyEvaluation", "mode=history, ticker, from_date=YYYY-MM-DD, include_composite=false",
 "Lịch sử 4 Key của VCB từ 01/2026",
 "Tóm tắt các mốc/ngày theo kết quả tool trả về."),

("Cau hoi ve danh gia 4 key co phieu.txt", "Lọc danh sách mã theo nhóm 4-key [date]",
 "Hỏi lấy danh sách/danh mục các mã theo nhóm 4-key (đúng sóng đúng ngành / đúng sóng sai ngành / sai sóng đúng ngành / sai sóng sai ngành)",
 "getStock4KeyScreen", "date=YYYY-MM-DD, group=dd/ds/sd/ss (không truyền ticker)",
 "Cung cấp danh sách các mã đúng sóng đúng ngành ngày 24/08/2026",
 "Trả về danh sách ticker ngắn gọn, cách nhau dấu phẩy. Nếu không có ngày → dùng ngày hiện tại."),

("Cau hoi ve danh gia 4 key co phieu.txt", "Đánh giá 4 key nhiều mã [date]",
 "Hỏi 4-key cho nhiều mã cụ thể cùng lúc",
 "getStock4KeyEvaluation", "mode=batch, tickers=[...], date=YYYY-MM-DD, include_composite=false",
 "Đánh giá 4 Key của SSI, VCB, HPG ngày 24/08/2026",
 "Trả lời dạng bảng ngắn: từng mã - nhóm 4-key - khuyến nghị."),

("Câu hỏi về giá của mã.txt", "Giá hiện tại của mã",
 "Hỏi giá cổ phiếu tại thời điểm hiện tại (hôm nay/hiện tại/bây giờ/latest/current)",
 "getTotalTradeReal", "ticker",
 "Giá SSI hiện tại là bao nhiêu?", ""),

("Câu hỏi về giá của mã.txt", "Giá tại một ngày cụ thể",
 "Hỏi giá cổ phiếu tại 1 ngày cụ thể",
 "getTotalTrade", "ticker, date",
 "Giá SSI ngày 20/08/2026 là bao nhiêu?", ""),

("Câu hỏi về lịch sử mua bán của một mã.txt", "Lịch sử mua bán [mã] giai đoạn YYYY-YYYY",
 "Hỏi lịch sử mua/bán của 1 mã trong nhiều năm",
 "getStockSignal", "ticker (sau đó tự lọc kết quả theo các năm được hỏi)",
 "Lịch sử mua bán SSI trong giai đoạn 2023-2025",
 "Trình bày tường thuật: ngày – tín hiệu – thay đổi tỷ trọng. Không suy diễn nếu API không trả dữ liệu."),

("Câu hỏi về mã, cổ phiếu, đạt chuẩn mã mạnh.txt", "Mã [X] bắt đầu mạnh từ khi nào?",
 "Hỏi mốc thời gian 1 mã bắt đầu/đạt chuẩn mã mạnh",
 "getSMDTTickerCross", "keyValue=ticker (không truyền date)",
 "Mã SSI bắt đầu mạnh từ khi nào?",
 "Lấy bản ghi có date MỚI NHẤT trong mảng smdts trả về, không lấy ngày đầu tiên."),

("Câu hỏi về mã, cổ phiếu, đạt chuẩn mã mạnh.txt", "Mã nào đạt chuẩn mã mạnh vào [ngày/tháng/năm]",
 "Hỏi danh sách mã đạt chuẩn mã mạnh tại 1 mốc thời gian",
 "getSMDTTickerCross", "date=yyyy-mm-dd hoặc yyyy-mm hoặc yyyy tuỳ độ chi tiết câu hỏi",
 "Mã nào đạt chuẩn mã mạnh vào tháng 07/2026?",
 "Nếu có nhiều mốc: mốc đầu ghi \"bắt đầu mạnh từ ngày...\", mốc sau ghi \"và tiếp tục đạt chuẩn mã mạnh lại từ ngày...\"."),

("Câu hỏi về mã, cổ phiếu, đạt chuẩn mã mạnh.txt", "Chỉ số của mã trong N phiên gần nhất",
 "Hỏi chỉ số/giá của 1 mã trong N phiên gần nhất",
 "getTotalTrade", "ticker, lastDates=N",
 "Giá SSI trong 20 phiên gần nhất là bao nhiêu?", "Liệt kê giá của tất cả các phiên trả về."),

("Câu hỏi về mã, cổ phiếu, đạt chuẩn mã mạnh.txt", "Lập bảng thống kê [Mã] theo tháng/năm/khoảng dài",
 "Hỏi thống kê 1 mã trải dài nhiều tháng hoặc nhiều năm",
 "getTotalTrade", "ticker; date=YYYY-MM (1 tháng) hoặc YYYY (1 năm); gọi LẶP LẠI nhiều lần nếu trải dài nhiều tháng/năm",
 "Lập bảng thống kê SSI từ 2022 đến nay",
 "Ví dụ cụ thể trong rule: gọi lần lượt date=2022, 2023, 2024, 2025 rồi ghép bảng. Không được bỏ sót năm nào."),

("Câu hỏi về mã, cổ phiếu, đạt chuẩn mã mạnh.txt", "[Ngày] cổ phiếu nào mạnh nhất dòng [ngành X]?",
 "Hỏi mã mạnh nhất trong 1 ngành tại 1 ngày",
 "getSMDTTicker", "date=ngày được hỏi; gọi lặp lại keyValue cho từng mã trong ngành",
 "Ngày 24/08/2026 cổ phiếu nào mạnh nhất dòng Ngân hàng?",
 "Mã có SMDT cao nhất trong ngành tại ngày đó là cổ phiếu mạnh nhất."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Ngành nào đang dẫn dắt",
 "Hỏi ngành nào đang dẫn dắt (có thể kèm hoặc không kèm ngày)",
 "getLeadingCoreBranches", "date=ngày được hỏi (bỏ trống nếu câu hỏi không có ngày)",
 "Ngành nào đang dẫn dắt hôm nay?",
 "Lấy TẤT CẢ ngành trả về, trả lời kèm SMDT của từng ngành."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Lộ trình các dòng dẫn sóng",
 "Hỏi lộ trình/diễn biến các dòng dẫn sóng theo thời gian (KHÔNG phải hỏi ngành nào mất vai trò)",
 "getCoreBranchLeader", "date=ngày/tháng được hỏi",
 "Lộ trình các dòng dẫn sóng trong tháng 08/2026?",
 "Lập bảng lộ trình dẫn sóng theo đúng dữ liệu trả về."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Ngành chủ lực nào dẫn sóng vào [date]",
 "Hỏi ngành chủ lực nào đang dẫn sóng tại 1 mốc thời gian cụ thể",
 "getSMDTBranchCross", "date=thời gian được hỏi",
 "Ngành chủ lực nào dẫn sóng vào 24/08/2026?",
 "Lọc kết quả trả về, chỉ giữ lại các ngành CHỦ LỰC, kèm SMDT tương ứng."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "[Ngành] mất vai trò dẫn sóng khi nào",
 "Hỏi mốc thời gian 1 ngành mất vai trò dẫn sóng",
 "getBranchPath rồi getSMDTBranchDrop", "Bước 1: getBranchPath(ticker) lấy path ngành. Bước 2: getSMDTBranchDrop(path), lấy lastDate",
 "Ngành Ngân hàng mất vai trò dẫn sóng khi nào?", ""),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Điều kiện ngành mạnh (kiến thức nghiệp vụ)",
 "Không phải câu hỏi gọi API — đây là NGƯỠNG nghiệp vụ dùng để diễn giải kết quả",
 "(không gọi API)", "—",
 "(áp dụng ngầm khi trả lời các câu hỏi về ngành mạnh)",
 "SMDT ngành trước đó <60%, sau vượt lên 60% nhưng vẫn <70% → có tiềm năng dẫn sóng/sắp tham gia dẫn sóng (áp dụng cho 6 ngành chủ lực)."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Sóng ngành chủ lực – ngưỡng \"đang dẫn sóng\" (kiến thức nghiệp vụ)",
 "Ngưỡng nghiệp vụ định nghĩa \"đang dẫn sóng\" vs \"đạt chuẩn ngành mạnh\"",
 "(không gọi API)", "—",
 "(áp dụng ngầm khi trả lời các câu hỏi về dẫn sóng)",
 "6 ngành chủ lực: \"dẫn sóng\" khi SMDT=70%, dòng tiền lan tỏa nhiều mã (không dồn vào vài mã), mã vốn hoá lớn cũng phải tích cực. Ngành KHÔNG chủ lực: \"đạt chuẩn ngành mạnh\" khi SMDT vượt 70%."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Ngành [X] bắt đầu dẫn sóng (6 ngành chủ lực) từ khi nào?",
 "Hỏi mốc CHỦ LỰC bắt đầu dẫn sóng/đạt chuẩn mạnh — chỉ lấy đúng 1 ngày gần nhất",
 "getSMDTBranchCross", "keyName=tên ngành",
 "Ngành Ngân hàng bắt đầu dẫn sóng từ khi nào?",
 "CHỈ lấy 1 ngày GẦN NHẤT so với hôm nay trong danh sách trả về, không nhắc các mốc cũ hơn. Khuôn mẫu bắt buộc: \"Ngành [X] bắt đầu dẫn sóng (đạt chuẩn ngành mạnh) từ ngày [ngày mới nhất].\" KHÔNG áp dụng cho câu hỏi dạng \"Sóng vào [date] ngành nào dẫn sóng?\"."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Thời điểm dòng đạt chuẩn ngành mạnh của 1 mã",
 "Hỏi mốc đạt chuẩn ngành mạnh nhưng nêu tên MÃ thay vì tên ngành — phải tự suy ra ngành của mã trước",
 "getSMDTBranchCross", "Bước 1: xác định ngành của mã. Bước 2: keyName=tên ngành, lấy ngày gần nhất",
 "SSI đạt chuẩn ngành mạnh từ khi nào? (SSI thuộc ngành Chứng khoán)",
 "Cùng khuôn mẫu trả lời như case \"Ngành [X] bắt đầu dẫn sóng\" ở trên."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Ngành chủ lực nào dẫn sóng hôm nay?",
 "Hỏi ngành chủ lực dẫn sóng tại phiên hiện tại (hôm nay/hiện tại), kể cả biến thể \"cho X mã dòng [ngành] còn mua được\"",
 "getSMDTBranch", "date=hôm nay",
 "Ngành chủ lực nào dẫn sóng hôm nay?",
 "Lọc trong 6 ngành chủ lực, giữ lại ngành có SMDT ≥ 70%."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Dòng nào dẫn sóng vào [tháng/năm quá khứ]",
 "Hỏi CHỦ LỰC dẫn sóng tại 1 mốc quá khứ (hôm qua/tháng/năm cụ thể)",
 "getSMDTBranchCross", "date=thời gian được hỏi",
 "Sóng tháng 07/2026 ngành nào dẫn sóng?",
 "Gọi xong lọc kết quả CHỈ GIỮ ngành chủ lực rồi trả lời."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Dòng nào đạt chuẩn ngành mạnh vào [tháng/năm quá khứ]",
 "Hỏi ngành KHÔNG CHỦ LỰC đạt chuẩn mạnh tại 1 mốc quá khứ",
 "getSMDTBranchCross", "date=thời gian được hỏi",
 "Ngành nào đạt chuẩn ngành mạnh vào năm 2026?",
 "Gọi xong lọc kết quả LOẠI BỎ các ngành chủ lực, chỉ giữ ngành còn lại — ngược với case phía trên."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Phân tích ngành [X]",
 "Hỏi \"phân tích ngành X\" — KHÔNG hiểu là tại thời điểm hiện tại",
 "(tuỳ ngữ cảnh, thường getSMDTBranch/getPerformance)", "Xem toàn bộ năm gần nhất có dữ liệu, không chỉ hôm nay",
 "Phân tích ngành Thép",
 "Lưu ý nghiệp vụ quan trọng: KHÔNG mặc định hiểu là hỏi hôm nay."),

("Câu hỏi về ngành, dẫn sóng, đạt chuẩn ngành mạnh.txt", "Sức mạnh dòng tiền ngành chủ lực vào [date] thế nào?",
 "Hỏi SMDT của TẤT CẢ 6 ngành chủ lực tại 1 ngày",
 "getSMDTBranch", "Gọi LẶP LẠI cho từng ngành trong 6 ngành chủ lực, date=ngày được hỏi",
 "Sức mạnh dòng tiền ngành chủ lực vào 24/08/2026 thế nào?",
 "Tổng hợp kết quả của cả 6 ngành rồi trả lời chung 1 bảng."),

("Câu hỏi về số lượng mua bán, chờ mua, chờ bán, độ tin cậy.txt", "Chờ mua/chờ bán/độ tin cậy tại [ngày]",
 "Hỏi số lượng mã tín hiệu mua/bán/chờ mua/chờ bán/tổng/độ tin cậy",
 "getStockWave", "date=ngày/tháng/năm được hỏi",
 "Chờ mua ngày 24/08/2026 là bao nhiêu?",
 "Trường trả về: buy, sell, waitbuy, waitsell, total, reliability(%)."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT các ngành chủ lực ngày [date]",
 "Hỏi SMDT của các ngành chủ lực tại 1 ngày",
 "getSMDTBranch", "path của từng ngành chủ lực, date=ngày được hỏi",
 "Sức mạnh dòng tiền các ngành chủ lực ngày 24/08/2026", "Lập bảng trả lời."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT ngành [X] từ [date] đến nay",
 "Hỏi SMDT 1 ngành trải dài từ 1 mốc tới hiện tại",
 "getSMDTBranch", "path ngành; gọi lặp lại theo từng tháng-năm từ mốc hỏi tới tháng-năm hiện tại (không truyền date đơn lẻ)",
 "SMDT ngành Ngân hàng từ tháng 01/2026 đến nay", "Lập bảng kết quả theo từng tháng."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT [ngành] là bao nhiêu?",
 "Hỏi SMDT của 1 ngành tại 1 thời điểm",
 "getSMDTBranch", "ngành, date được hỏi",
 "SMDT ngành Thép là bao nhiêu?", "Trả lời kèm ký hiệu %."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT [mã cổ phiếu] là bao nhiêu?",
 "Hỏi SMDT — cần phân biệt mã cổ phiếu thật (SSI, VCB, HPG...) hay tên ngành",
 "getSMDTTicker (nếu là mã thật) / getSMDTBranch (nếu là tên ngành)",
 "Nếu giá trị là ngành (ngân hàng, chứng khoán, thép...) → BẮT BUỘC gọi getSMDTBranch, KHÔNG được gọi getSMDTTicker",
 "SMDT SSI là bao nhiêu?",
 "Lỗi thường gặp cần tránh: gọi nhầm getSMDTTicker khi user thực ra đang hỏi tên ngành."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "Mã giá giảm hôm nay mà SMDT tăng dần đều",
 "Hỏi lập bảng thống kê mã giảm giá nhưng SMDT vẫn tăng đều",
 "getTickersPriceDownSMDTIncreasing", "không truyền date",
 "Lập bảng thống kê mã giá giảm hôm nay mà SMDT tăng dần đều",
 "Trả lời TOÀN BỘ mã trả về, không được rút gọn bằng dấu \"...\"."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT cổ phiếu [X] từ tháng [month] đến nay",
 "Hỏi SMDT 1 mã trải dài từ 1 tháng tới hiện tại",
 "getSMDTTicker", "ticker; gọi lặp lại theo từng tháng cho tới tháng hiện tại",
 "SMDT SSI từ tháng 01/2026 đến nay", ""),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT các mã dòng [ngành] của [ngày]",
 "Hỏi SMDT của TẤT CẢ mã trong 1 ngành tại 1 ngày",
 "getBranchPath rồi getSMDTTicker", "Bước 1: getBranchPath(ngành) lấy list mã. Bước 2: getSMDTTicker từng mã, date=ngày hỏi",
 "SMDT các mã dòng Ngân hàng của ngày 24/08/2026", "Tổng hợp kết quả rồi trả lời."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT các mã dòng [ngành] từ tháng [month] đến nay",
 "Hỏi SMDT toàn bộ mã trong 1 ngành trải dài thời gian",
 "getBranchSMDTTickers", "ngành, date được hỏi",
 "SMDT các mã dòng Ngân hàng từ tháng 01/2026 đến nay", ""),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "SMDT của [mã/ngành] trong X phiên vừa qua",
 "Hỏi SMDT N phiên gần nhất của 1 mã hoặc 1 ngành",
 "getSMDTLastN", "n=X; ticker (nếu là mã) hoặc keyName/branch_path (nếu là ngành)",
 "SMDT SSI trong 10 phiên vừa qua", ""),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "Mã có SMDT tăng dần đều vào [date]",
 "Hỏi danh sách mã có SMDT tăng dần đều tại 1 ngày",
 "getSMDTIncreasing3", "date=ngày được hỏi (bỏ trống nếu không nói ngày; bỏ trống ticker nếu không nói mã)",
 "Các mã có SMDT tăng dần đều vào 24/08/2026", "Lập bảng."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "Lộ trình các mã suy yếu trong khoảng [date]",
 "Hỏi các mã suy yếu (SMDT giảm) trong 1 khoảng thời gian",
 "getSMDTTickerDrop", "date được hỏi, hoặc dateFrom & dateTo",
 "Lộ trình các mã suy yếu trong năm 2026", "Lập bảng."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "Lộ trình các dòng chủ lực suy yếu trong [date]",
 "Hỏi các ngành chủ lực suy yếu trong 1 khoảng thời gian",
 "getSMDTBranchDrop", "date được hỏi, hoặc dateFrom & dateTo",
 "Lộ trình các dòng chủ lực suy yếu trong tháng 07/2026", "Lập bảng."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "Ngành có nhiều mã SMDT tăng dần đều nhất vào [date]",
 "Hỏi ngành dẫn đầu về số mã có SMDT tăng dần đều",
 "getTopBranchSMDTIncreasing", "date được hỏi (bỏ trống nếu không có ngày)",
 "Ngành nào có số mã có SMDT tăng dần đều nhiều nhất vào 24/08/2026?", "Lập bảng."),

("Câu hỏi về sức mạnh dòng tiền, smdt ngành, mã.txt", "Danh sách mã vượt 70% kèm giá của [ngành] vào [date]",
 "Hỏi danh sách mã trong ngành có SMDT vượt 70%, kèm giá",
 "getBranchStrongSMDTWithPrice", "keyName hoặc path của ngành, date được hỏi",
 "Danh sách các mã vượt 70% kèm giá của ngành Ngân hàng vào 24/08/2026", "Lập bảng."),

("Câu hỏi về tín hiệu dòng tiền mã , dòng tiền mã.txt", "Dòng tiền bắt đầu đổ vào tháng X khi nào (ngành/mã)",
 "QUAN TRỌNG: đây là DÒNG TIỀN, KHÔNG PHẢI SMDT — tuyệt đối không gọi API SMDT hay suy luận qua %SMDT",
 "getCashFlowBranch (ngành) / getCashFlowTicker (mã)", "path (ngành) hoặc ticker (mã), lọc theo tháng X được hỏi, lấy DATA ĐẦU TIÊN của tháng đó",
 "Dòng tiền ngành Ngân hàng bắt đầu đổ vào tháng 07/2026 khi nào?",
 "Lỗi thường gặp cần tránh: nhầm sang hỏi SMDT ngành/mã."),

("Câu hỏi về tín hiệu dòng tiền mã , dòng tiền mã.txt", "Dòng tiền [ngành] hiện nay thế nào?",
 "Hỏi dòng tiền ngành không nói rõ ngày",
 "getCashFlowBranch", "path=path ngành, date=hôm nay; nếu hôm nay không có data thì gọi lại KHÔNG truyền date, lấy ngày gần nhất có data",
 "Dòng tiền ngành Ngân hàng hiện nay thế nào?", ""),

("Câu hỏi về tín hiệu dòng tiền mã , dòng tiền mã.txt", "Dòng tiền [mã] hiện nay thế nào?",
 "Hỏi dòng tiền của 1 mã không nói rõ ngày",
 "getCashFlowTicker", "ticker, date=hôm nay; nếu hôm nay không có data thì gọi lại KHÔNG truyền date, lấy ngày gần nhất có data",
 "Dòng tiền SSI hiện nay thế nào?", ""),

("Câu hỏi về tín hiệu dòng tiền mã , dòng tiền mã.txt", "Dòng tiền [mã/ngành] là bao nhiêu?",
 "Hỏi dòng tiền tại 1 ngày cụ thể",
 "getCashFlowTicker (mã) / getCashFlowBranch (ngành)", "ticker hoặc ngành, date được hỏi",
 "Dòng tiền SSI ngày 24/08/2026 là bao nhiêu?", ""),

("Câu hỏi về tín hiệu dòng tiền mã , dòng tiền mã.txt", "Tín hiệu dòng tiền [ticker] từ [date] đến nay",
 "Hỏi dòng tiền 1 mã trải dài từ 1 mốc tới hiện tại",
 "getCashFlowTicker", "gọi lặp lại theo từng tháng-năm từ mốc hỏi tới tháng-năm hiện tại",
 "Tín hiệu dòng tiền ACB từ 4/5 đến nay (ví dụ gốc trong rule, ngày hiện tại giả định 2026-06-12)",
 "Ví dụ đúng nguyên văn trong rule: gọi getCashFlowTicker(ticker=ACB, date=2026-05) rồi (date=2026-06). Trình bày text thường, không markdown/bold."),

("Câu hỏi về tín hiệu giao dịch (mua,bán), giá vốn trung bình, tỷ trọng nắm giữ, tỷ trọng giao dịch của mã.txt",
 "Tín hiệu mua/bán, giá vốn TB, tỷ trọng nắm giữ/giao dịch của mã",
 "Hỏi các chỉ số giao dịch của 1 mã",
 "getStockSignal", "ticker, lấy bản ghi lastdate",
 "Giá vốn trung bình và tỷ trọng nắm giữ của SSI hiện tại?",
 "Trường trả về: price, ave, hold, percent, smdt, trade(1=mua,2=bán)."),

("Câu hỏi về tín hiệu giao dịch (mua,bán), giá vốn trung bình, tỷ trọng nắm giữ, tỷ trọng giao dịch của mã.txt",
 "Tín hiệu mua/bán gần nhất của [mã]",
 "Hỏi tín hiệu giao dịch gần nhất",
 "getStockSignal", "ticker, lấy tín hiệu cuối cùng",
 "Tín hiệu mua bán gần nhất của SSI?", ""),

("Câu hỏi về tín hiệu giao dịch (mua,bán), giá vốn trung bình, tỷ trọng nắm giữ, tỷ trọng giao dịch của mã.txt",
 "Chỉ số của mã [X] trong [N] phiên gần nhất",
 "Hỏi chỉ số giao dịch N phiên gần nhất",
 "getTotalTrade", "ticker, lastDates=N",
 "Chỉ số của SSI trong 15 phiên gần nhất?", ""),

("Câu hỏi về xác nhận chân sóng, [tháng, năm] là sóng lớn hay sóng hồi.txt",
 "[DD/MM] có xác nhận chân sóng không? / sóng DD-MM là sóng lớn hay sóng hồi",
 "Hỏi xác nhận chân sóng hoặc phân loại sóng lớn/sóng hồi tại 1 ngày",
 "getAnalyzeWave", "date=ngày được hỏi",
 "28/07 là sóng lớn hay sóng hồi?",
 "BẮT BUỘC trả lời CHÍNH XÁC 100% nội dung API trả về, không tự diễn giải định nghĩa chung chung."),

("Câu hỏi về xác nhận chân sóng, [tháng, năm] là sóng lớn hay sóng hồi.txt",
 "Chân sóng gần nhất là ngày nào?",
 "Hỏi mốc chân sóng gần nhất",
 "getChanSong", "không truyền tham số",
 "Chân sóng gần nhất là ngày nào?",
 "Ưu tiên bản ghi mới nhất có tín hiệu \"Xác nhận tạo đáy\" (hoặc \"Chuẩn bị tạo đáy\" nếu hỏi giai đoạn chuẩn bị)."),

("Câu hỏi về xác nhận chân sóng, [tháng, năm] là sóng lớn hay sóng hồi.txt",
 "Phiên chuẩn bị/xác nhận tạo đáy gần nhất là ngày nào?",
 "Hỏi phiên gần nhất theo đúng loại tín hiệu được nêu",
 "getChanSong", "không truyền tham số",
 "Phiên xác nhận tạo đáy gần nhất là ngày nào?",
 "Trả lời ngắn gọn: Ngày, Tín hiệu, Chờ mua, Mua. Không nhắc tên API."),

("Câu hỏi về xác nhận chân sóng, [tháng, năm] là sóng lớn hay sóng hồi.txt",
 "Trong [tháng/năm] có những phiên chuẩn bị/xác nhận tạo đáy nào?",
 "Hỏi liệt kê các phiên trong 1 khoảng thời gian",
 "getChanSong", "không truyền tham số, tự lọc theo tháng/năm và tín hiệu được hỏi",
 "Trong tháng 07/2026 có những phiên xác nhận tạo đáy nào?",
 "Liệt kê theo thời gian: Ngày, Tín hiệu, Chờ mua, Mua."),

("Hiệu suất cổ phiếu khi dẫn sóng.txt", "Hiệu suất cổ phiếu [X] khi ngành dẫn sóng/đạt chuẩn ngành mạnh",
 "Hỏi hiệu suất tăng giá của 1 mã gắn với thời điểm ngành nó thuộc về dẫn sóng",
 "getPerformance", "Bước 1: xác định path ngành của mã. Bước 2: branch_path=path",
 "Hiệu suất cổ phiếu SSI khi ngành dẫn sóng?",
 "Trình bày: ngành thuộc về, ngày ngành bắt đầu dẫn sóng, ngày/giá đáy, ngày/giá đỉnh, hiệu suất tăng giá. Không show công thức/thuật ngữ kỹ thuật (zigzag...)."),

("Hiệu suất cổ phiếu khi dẫn sóng.txt", "Hiệu suất các cổ phiếu [ngành] vào tháng MM-YYYY khi ngành dẫn sóng",
 "Hỏi hiệu suất toàn bộ mã trong 1 ngành, tại 1 tháng cụ thể, khi ngành đó dẫn sóng",
 "getBranchPath rồi getPerformance", "Bước 1: getBranchPath(name=ngành) lấy path (KHÔNG truyền date). Bước 2: getPerformance(branch_path=path, date=mm-yyyy)",
 "Hiệu suất các cổ phiếu ngành Ngân hàng vào tháng 07/2026 khi ngành dẫn sóng?",
 "Ví dụ tham số đúng trong rule: branch_path=7-211-212-213-214-&date=07-2025. TUYỆT ĐỐI không gọi getPerformance bằng branch=[tên ngành] — chỉ được dùng branch_path, nếu không sẽ thiếu/sai dữ liệu."),
]

wb = Workbook()
ws = wb.active
ws.title = "Case Catalog"

headers = ["STT", "File rule nguồn", "Tên Guide / Case", "Mô tả / Khi nào áp dụng",
           "API gọi", "Tham số", "Câu hỏi ví dụ cụ thể", "Ghi chú"]
ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER
ws.freeze_panes = "A2"

for i, row in enumerate(ROWS, start=1):
    r = ws.append((i,) + row)

for row_idx in range(2, len(ROWS) + 2):
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.font = CELL_FONT
        c.alignment = WRAP
        c.border = BORDER

widths = [5, 30, 34, 42, 22, 40, 40, 46]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row_idx in range(2, len(ROWS) + 2):
    ws.row_dimensions[row_idx].height = 60

out_path = "case_catalog.xlsx"
wb.save(out_path)
print("Saved", out_path, "rows:", len(ROWS))
